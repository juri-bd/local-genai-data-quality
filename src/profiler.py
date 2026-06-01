import pandas as pd
import ollama
import json
import sys
import io
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -------------------------------------------------------------------
# CHECKS
# -------------------------------------------------------------------

def check_nulls(df):
    null_counts = df.isnull().sum()
    return {
        col: {"null_count": int(c), "null_pct": round(c / len(df) * 100, 1)}
        for col, c in null_counts.items() if c > 0
    }

def check_duplicates(df):
    return {"duplicate_rows": int(df.duplicated().sum())}

def check_dtypes(df):
    return {col: str(dtype) for col, dtype in df.dtypes.items()}

def check_numeric_ranges(df):
    result = {}
    for col in df.select_dtypes(include="number").columns:
        result[col] = {
            "min": float(df[col].min()),
            "max": float(df[col].max()),
            "mean": round(float(df[col].mean()), 2),
            "negatives": int((df[col] < 0).sum()),
            "zeros": int((df[col] == 0).sum()),
            "p25": float(df[col].quantile(0.25)),
            "p75": float(df[col].quantile(0.75)),
        }
    return result

def check_potential_key_duplicates(df):
    result = {}
    for col in df.select_dtypes(include="object").columns:
        dupes = df[col].duplicated().sum()
        if dupes > 0:
            result[col] = {"duplicate_values": int(dupes)}
    return result

def check_value_distributions(df):
    result = {}
    for col in df.select_dtypes(include="object").columns:
        counts = df[col].value_counts(normalize=True).head(5)
        result[col] = {str(k): f"{v*100:.1f}%" for k, v in counts.items()}
        if df[col].dropna().str.lower().nunique() < df[col].dropna().nunique():
            result[col]["_warning"] = "inconsistent capitalisation detected"
    return result

def check_constant_columns(df):
    result = {}
    for col in df.columns:
        if df[col].nunique() <= 1:
            result[col] = {"unique_values": int(df[col].nunique()), "issue": "constant or empty column"}
    return result

CHECKS = [
    ("nulls", check_nulls),
    ("duplicates", check_duplicates),
    ("dtypes", check_dtypes),
    ("numeric_ranges", check_numeric_ranges),
    ("potential_key_duplicates", check_potential_key_duplicates),
    ("value_distributions", check_value_distributions),
    ("constant_columns", check_constant_columns),
]


# -------------------------------------------------------------------
# Instant programmatic dashboard
# -------------------------------------------------------------------

def quick_audit(df) -> pd.DataFrame:
    issues = []

    for col in df.columns:
        null_pct = df[col].isnull().mean() * 100
        if null_pct > 50:
            issues.append({"column": col, "issue": f"{null_pct:.1f}% null values", "severity": "critical"})
        elif null_pct > 10:
            issues.append({"column": col, "issue": f"{null_pct:.1f}% null values", "severity": "warning"})
        elif null_pct > 0:
            issues.append({"column": col, "issue": f"{null_pct:.1f}% null values", "severity": "info"})

    dupe_count = int(df.duplicated().sum())
    if dupe_count > 0:
        sev = "critical" if dupe_count > len(df) * 0.01 else "warning"
        issues.append({"column": "all", "issue": f"{dupe_count} duplicate rows", "severity": sev})

    for col in df.select_dtypes(include="number").columns:
        q1 = df[col].quantile(0.25)
        q3 = df[col].quantile(0.75)
        iqr = q3 - q1
        outlier_count = int(((df[col] < q1 - 3 * iqr) | (df[col] > q3 + 3 * iqr)).sum())
        if outlier_count > 0:
            issues.append({"column": col, "issue": f"{outlier_count} extreme outliers (3x IQR)", "severity": "warning"})
        neg_count = int((df[col] < 0).sum())
        if neg_count > 0 and df[col].mean() > 0:
            issues.append({"column": col, "issue": f"{neg_count} negative values", "severity": "warning"})
        zero_count = int((df[col] == 0).sum())
        if zero_count > len(df) * 0.1:
            issues.append({"column": col, "issue": f"{zero_count} zero values ({zero_count/len(df)*100:.1f}%)", "severity": "info"})
        if df[col].nunique() == 1:
            issues.append({"column": col, "issue": "constant column, all values identical", "severity": "critical"})

    for col in df.select_dtypes(include="object").columns:
        if df[col].nunique() == len(df):
            issues.append({"column": col, "issue": "all values unique, possible ID column", "severity": "info"})
        if df[col].dropna().str.lower().nunique() < df[col].dropna().nunique():
            issues.append({"column": col, "issue": "inconsistent capitalisation", "severity": "warning"})
        if df[col].nunique() == 1:
            issues.append({"column": col, "issue": "constant column, all values identical", "severity": "critical"})

    return pd.DataFrame(issues) if issues else pd.DataFrame(columns=["column", "issue", "severity"])


# -------------------------------------------------------------------
# Suspicious row extraction
# -------------------------------------------------------------------

def extract_suspicious_rows(df: pd.DataFrame, max_rows: int = 100) -> pd.DataFrame:
    suspicious_idx = set()

    for col in df.columns:
        null_rows = df[df[col].isnull()].index.tolist()
        suspicious_idx.update(null_rows[:3])

    dupe_rows = df[df.duplicated(keep=False)].index.tolist()
    suspicious_idx.update(dupe_rows[:20])

    for col in df.select_dtypes(include="number").columns:
        q1 = df[col].quantile(0.25)
        q3 = df[col].quantile(0.75)
        iqr = q3 - q1
        outliers = df[(df[col] < q1 - 3 * iqr) | (df[col] > q3 + 3 * iqr)].index.tolist()
        suspicious_idx.update(outliers[:10])
        if df[col].mean() > 0:
            negs = df[df[col] < 0].index.tolist()
            suspicious_idx.update(negs[:10])

    for col in df.select_dtypes(include="object").columns:
        counts = df[col].value_counts()
        rare = counts[counts <= 2].index
        rare_rows = df[df[col].isin(rare)].index.tolist()
        suspicious_idx.update(rare_rows[:10])

    return df.loc[list(suspicious_idx)].head(max_rows)


# -------------------------------------------------------------------
# Ollama
# -------------------------------------------------------------------

def build_prompt(profile: dict, df=None) -> str:
    sample = ""
    if df is not None:
        suspicious = extract_suspicious_rows(df)
        null_summary = {
            col: f"{round(df[col].isnull().mean()*100, 1)}% null"
            for col in df.columns if df[col].isnull().any()
        }
        sample = (
            f"Total rows: {len(df)}.\n"
            f"Null summary: {json.dumps(null_summary, separators=(',', ':'))}\n"
            f"Suspicious rows ({len(suspicious)} sampled):\n"
            + suspicious.to_csv(index=True)
        )
    compact_profile = json.dumps(profile, separators=(',', ':'))
    return (
        "You are a data quality auditor. Find data quality problems.\n\n"
        "OUTPUT: max 10 bullet points, one line each, no intro, no explanation.\n\n"
        "LOOK FOR: nulls, duplicates, negatives where impossible, "
        "zeros where suspicious, inconsistent formatting, outliers, wrong types, "
        "constant columns, capitalisation inconsistencies.\n\n"
        f"PROFILE:{compact_profile}\n\n"
        f"{sample}\n\n"
        "Issues found:"
    )

def ask_ollama_streaming(profile: dict, df=None, model: str = "gemma3:4b"):
    prompt = build_prompt(profile, df)
    stream = ollama.chat(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        stream=True
    )
    for chunk in stream:
        yield chunk["message"]["content"]

def ask_ollama(profile: dict, df=None, model: str = "gemma3:4b") -> str:
    return "".join(ask_ollama_streaming(profile, df, model))

def ask_ollama_rules(profile: dict, df=None, model: str = "gemma3:4b") -> list:
    sample = ""
    if df is not None:
        suspicious = extract_suspicious_rows(df)
        null_summary = {
            col: f"{round(df[col].isnull().mean()*100, 1)}% null"
            for col in df.columns if df[col].isnull().any()
        }
        sample = (
            f"Total rows: {len(df)}.\n"
            f"Null summary: {json.dumps(null_summary, separators=(',', ':'))}\n"
            f"Suspicious rows ({len(suspicious)} sampled):\n"
            + suspicious.to_csv(index=True)
        )
    compact_profile = json.dumps(profile, separators=(',', ':'))
    prompt = (
        "You are a data quality auditor. Analyze this dataset and return ONLY a JSON array of rules to validate.\n\n"
        "Each rule must follow this exact format:\n"
        '{"column": "column_name", "description": "human readable issue", '
        '"pandas_filter": "python expression using df[column] that returns True for BAD rows"}\n\n'
        "Examples:\n"
        '{"column": "age", "description": "age must be between 0 and 120", "pandas_filter": "(df[\'age\'] < 0) | (df[\'age\'] > 120)"}\n'
        '{"column": "loan_amount", "description": "loan amount must be positive", "pandas_filter": "df[\'loan_amount\'] <= 0"}\n\n'
        "Return ONLY the JSON array. No explanation, no markdown, no backticks.\n\n"
        f"PROFILE:{compact_profile}\n\n"
        f"{sample}\n\nRules:"
    )
    response = ollama.chat(model=model, messages=[{"role": "user", "content": prompt}])
    raw = response["message"]["content"].strip()
    try:
        raw = raw.replace("```json", "").replace("```", "").strip()
        return json.loads(raw)
    except json.JSONDecodeError:
        return []

def validate_rules(df: pd.DataFrame, rules: list) -> pd.DataFrame:
    results = []
    for rule in rules:
        try:
            mask = eval(rule["pandas_filter"], {"df": df})
            failing_rows = int(mask.sum())
            results.append({
                "column": rule["column"],
                "description": rule["description"],
                "failing_rows": failing_rows,
                "total_rows": len(df),
                "fail_pct": round(failing_rows / len(df) * 100, 1),
                "passed": failing_rows == 0,
                "filter": rule["pandas_filter"]
            })
        except Exception as e:
            results.append({
                "column": rule["column"],
                "description": rule["description"],
                "failing_rows": -1,
                "total_rows": len(df),
                "fail_pct": -1,
                "passed": False,
                "filter": f"ERROR: {e}"
            })
    return pd.DataFrame(results)

def ask_ollama_batched(profile: dict, df: pd.DataFrame, model: str = "gemma3:4b",
                       batch_size: int = 100, progress_callback=None) -> str:
    suspicious = extract_suspicious_rows(df, max_rows=500)
    batches = [suspicious[i:i+batch_size] for i in range(0, len(suspicious), batch_size)]
    all_findings = [None] * len(batches)
    completed = 0

    def process_batch(args):
        i, batch = args
        prompt = (
            f"Data quality auditor. Batch {i+1}/{len(batches)} of suspicious rows "
            f"from a {len(df)}-row dataset.\n"
            "Bullet points only, one line each, no intro.\n\n"
            f"{batch.to_csv(index=True)}\n\nIssues:"
        )
        response = ollama.chat(model=model, messages=[{"role": "user", "content": prompt}])
        return i, response["message"]["content"]

    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(process_batch, (i, batch)): i
                   for i, batch in enumerate(batches)}
        for future in as_completed(futures):
            i, result = future.result()
            all_findings[i] = result
            completed += 1
            if progress_callback:
                progress_callback(completed, len(batches))

    consolidation_prompt = (
        f"Data quality expert. Dataset had {len(df)} rows. "
        f"Profile:{json.dumps(profile, separators=(',', ':'))}\n\n"
        "Batch findings:\n"
        + "\n---\n".join(all_findings)
        + "\n\nConsolidate into max 10 unique prioritized bullet points:"
    )
    response = ollama.chat(model=model, messages=[{"role": "user", "content": consolidation_prompt}])
    return response["message"]["content"]


# -------------------------------------------------------------------
# Export report
# -------------------------------------------------------------------

def _style_header(ws, row, cols, bg_color="1F3864", font_color="FFFFFF"):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=font_color, name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=bg_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _style_severity_cell(cell, severity):
    colors = {
        "critical": "FF4B4B",
        "warning":  "FFA500",
        "info":     "4B8BFF",
    }
    color = colors.get(severity, "CCCCCC")
    cell.fill = PatternFill("solid", start_color=color)
    cell.font = Font(color="FFFFFF", name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="center")

def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

def _write_df(ws, df, start_row=2, header_color="1F3864"):
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=header_color)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value).font = Font(name="Arial", size=10)

def build_report(
    filename: str,
    df: pd.DataFrame,
    audit_df: pd.DataFrame,
    ai_analysis: str = None,
    rules_df: pd.DataFrame = None,
    failing_rows: dict = None,
) -> bytes:
    wb = Workbook()

    # ---- Sheet 1: Overview ----
    ws1 = wb.active
    ws1.title = "Overview"
    ws1["A1"] = "Data Quality Report"
    ws1["A1"].font = Font(bold=True, size=16, color="1F3864", name="Arial")
    ws1["A2"] = f"File: {filename}"
    ws1["A3"] = f"Rows: {len(df)}"
    ws1["A4"] = f"Columns: {len(df.columns)}"
    ws1["A5"] = f"Column names: {', '.join(df.columns.tolist())}"
    for row in [2, 3, 4, 5]:
        ws1.cell(row=row, column=1).font = Font(name="Arial", size=11)
    ws1.column_dimensions["A"].width = 80

    # ---- Sheet 2: Instant Audit ----
    ws2 = wb.create_sheet("Instant Audit")
    ws2["A1"] = "Programmatic Findings"
    ws2["A1"].font = Font(bold=True, size=13, color="1F3864", name="Arial")

    if not audit_df.empty:
        _write_df(ws2, audit_df, start_row=2)
        # color severity column
        sev_col = list(audit_df.columns).index("severity") + 1
        for row_idx, sev in enumerate(audit_df["severity"], 3):
            _style_severity_cell(ws2.cell(row=row_idx, column=sev_col), sev)
    else:
        ws2["A3"] = "No issues detected."
    _autofit(ws2)

    # ---- Sheet 3: AI Analysis ----
    ws3 = wb.create_sheet("AI Analysis")
    ws3["A1"] = "AI-Generated Analysis"
    ws3["A1"].font = Font(bold=True, size=13, color="1F3864", name="Arial")
    if ai_analysis:
        for i, line in enumerate(ai_analysis.split("\n"), 3):
            ws3.cell(row=i, column=1, value=line).font = Font(name="Arial", size=10)
    else:
        ws3["A3"] = "No AI analysis run yet."
    ws3.column_dimensions["A"].width = 100

    # ---- Sheet 4: Rule Validation ----
    ws4 = wb.create_sheet("Rule Validation")
    ws4["A1"] = "Rule Validation Results"
    ws4["A1"].font = Font(bold=True, size=13, color="1F3864", name="Arial")
    if rules_df is not None and not rules_df.empty:
        display_cols = ["column", "description", "failing_rows", "fail_pct", "passed"]
        _write_df(ws4, rules_df[display_cols], start_row=2)
        passed_col = display_cols.index("passed") + 1
        for row_idx, passed in enumerate(rules_df["passed"], 3):
            cell = ws4.cell(row=row_idx, column=passed_col)
            if passed:
                cell.fill = PatternFill("solid", start_color="2ECC71")
                cell.font = Font(color="FFFFFF", name="Arial", size=10, bold=True)
            else:
                cell.fill = PatternFill("solid", start_color="FF4B4B")
                cell.font = Font(color="FFFFFF", name="Arial", size=10, bold=True)
    else:
        ws4["A3"] = "No rule validation run yet."
    _autofit(ws4)

    # ---- Sheet 5+: Failing rows per rule ----
    if failing_rows:
        for rule_desc, fdf in failing_rows.items():
            safe_name = rule_desc[:28].replace("/", "-").replace("\\", "-").replace(":", "-")
            ws = wb.create_sheet(f"Fail - {safe_name}")
            ws["A1"] = f"Failing rows: {rule_desc}"
            ws["A1"].font = Font(bold=True, size=11, color="1F3864", name="Arial")
            if not fdf.empty:
                _write_df(ws, fdf.head(50), start_row=2)
            _autofit(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# -------------------------------------------------------------------
# CLI entry point
# -------------------------------------------------------------------

def run(csv_path: str):
    print(f"Loading {csv_path}...")
    df = pd.read_csv(csv_path)
    print(f"{len(df)} rows, {len(df.columns)} columns.\n")

    profile = {"row_count": len(df), "column_count": len(df.columns)}
    for name, check in CHECKS:
        profile[name] = check(df)

    print("=== Instant Audit ===")
    audit = quick_audit(df)
    print(audit.to_string(index=False))

    print("\n=== AI Analysis ===")
    for chunk in ask_ollama_streaming(profile, df):
        print(chunk, end="", flush=True)
    print()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python src/profiler.py path/to/dataset.csv")
        sys.exit(1)
    run(sys.argv[1])